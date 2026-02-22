# -*- coding: utf-8 -*-
# app.py – Console d’envoi des échéances (PDF + Excel) avec boutons (Streamlit)

import os
import re
import time
import json
import ssl
import uuid
import csv
from urllib.parse import quote_plus
import smtplib
import mimetypes
import streamlit as st
import pandas as pd
from math import ceil
from datetime import datetime
from email.message import EmailMessage
from email import policy
from email.parser import BytesParser


def _format_email_html(body: str) -> str:
    parts = [p.strip() for p in re.split(r"\n\s*\n", body.strip()) if p.strip()]
    html_parts = []
    for p in parts:
        p = p.replace("\n", "<br>")
        html_parts.append(f"<p style=\"margin:0 0 12px 0;\">{p}</p>")
    return "\n".join(html_parts)

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# === Helper PDF (ton helper.py compact premium) ===
from helper import Helper  # doit être dans le même dossier
helper = Helper()

# ---------------------- Utilitaires Excel / SMTP / Email ----------------------

def style_excel(xlsx_path: str, sheet_name: str = "Contrats", date_col_name: str = "DATE ECHEANCE"):
    """Applique un style simple et lisible sur le fichier Excel."""
    wb = load_workbook(xlsx_path)
    ws = wb[sheet_name]

    max_row, max_col = ws.max_row, ws.max_column

    header_fill = PatternFill(start_color="FF8C00", end_color="FF8C00", fill_type="solid")
    header_font = Font(name="Helvetica", bold=True, color="FFFFFF", size=10)
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for j in range(1, max_col + 1):
        c = ws.cell(row=1, column=j)
        c.fill, c.font, c.alignment = header_fill, header_font, header_align

    thin = Side(style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    body_font = Font(name="Helvetica", size=9)
    body_align = Alignment(horizontal="left", vertical="top", wrap_text=True)

    for i in range(2, max_row + 1):
        for j in range(1, max_col + 1):
            c = ws.cell(row=i, column=j)
            c.border, c.font, c.alignment = border, body_font, body_align

    # largeur auto basique (en fonction de l'entête)
    for j in range(1, max_col + 1):
        header_text = str(ws.cell(1, j).value or "")
        base = max(12, min(60, int(len(header_text) * 1.2) + 2))
        ws.column_dimensions[get_column_letter(j)].width = base

    # format date si présent
    try:
        date_idx = next((j for j in range(1, max_col + 1)
                         if str(ws.cell(1, j).value).strip().upper() == date_col_name.upper()), None)
        if date_idx:
            for r in range(2, max_row + 1):
                ws.cell(r, date_idx).number_format = "DD/MM/YYYY"
    except Exception:
        pass

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    wb.save(xlsx_path)


def open_smtp(host, port, user, pwd, log=print):
    """Ouvre une connexion SMTP (Office365) avec TLS."""
    ctx = ssl.create_default_context()
    server = smtplib.SMTP(host, port, timeout=60)
    server.ehlo()
    server.starttls(context=ctx)
    server.ehlo()
    server.login(user, pwd)
    log("📡 Connexion SMTP OK")
    return server


def send_email_with_attachments(
    server,
    from_email,
    to_email,
    subject,
    body,
    attachments: list,
    dry_run: bool,
    test_to: str,
    request_receipt: bool = True,
    envelope_from: str | None = None,
    report_id: str | None = None,
):
    msg = EmailMessage()
    msg["Subject"] = subject
    msg["From"]    = f"Leadway Renouvellements <{from_email}>"
    msg["To"]      = test_to if dry_run else to_email
    msg["Reply-To"]= from_email
    if report_id:
        msg["X-Report-ID"] = report_id
    if request_receipt:
        msg["Disposition-Notification-To"] = from_email
        msg["Return-Receipt-To"] = from_email
    # Corps email: texte + HTML si des balises sont présentes
    if "<" in body and ">" in body:
        plain_body = re.sub(r"<[^>]+>", "", body)
        msg.set_content(plain_body)
        html_body = _format_email_html(body)
        msg.add_alternative(html_body, subtype="html")
    else:
        msg.set_content(body)
        msg.add_alternative(_format_email_html(body), subtype="html")

    for path in attachments:
        if not os.path.exists(path):
            st.warning(f"PJ introuvable : {path}")
            continue
        ctype,_ = mimetypes.guess_type(path)
        if ctype is None: ctype = "application/octet-stream"
        maintype,subtype = ctype.split("/",1)
        with open(path,"rb") as f:
            msg.add_attachment(f.read(), maintype=maintype, subtype=subtype,
                               filename=os.path.basename(path))

    mail_opts = ["RET=HDRS"] if request_receipt else []
    rcpt_opts = ["NOTIFY=SUCCESS,FAILURE,DELAY"] if request_receipt else []
    return server.send_message(
        msg,
        from_addr=envelope_from or from_email,
        to_addrs=[test_to if dry_run else to_email],
        mail_options=mail_opts,
        rcpt_options=rcpt_opts,
    )



def parse_dsn_eml(eml_bytes: bytes) -> list[dict]:
    msg = BytesParser(policy=policy.default).parsebytes(eml_bytes)
    results = []
    for part in msg.walk():
        ctype = part.get_content_type()
        if ctype == "message/delivery-status":
            payload = part.get_payload()
            for item in payload:
                results.append({
                    "Final-Recipient": item.get("Final-Recipient", ""),
                    "Action": item.get("Action", ""),
                    "Status": item.get("Status", ""),
                    "Diagnostic-Code": item.get("Diagnostic-Code", ""),
                    "Original-Message-ID": item.get("Original-Message-ID", ""),
                })
    return results


# ------------------------------- UI Streamlit ---------------------------------

st.set_page_config(page_title="Envoi Échéances Leadway", page_icon="📨", layout="wide")

st.title("📨 Console d’envoi des échéances – Leadway")
st.caption("Génération PDF/Excel par agent et envoi par vagues via SMTP.")

# ---- Confirmation de lecture (via lien) ----
params = st.query_params
if params.get("confirm") == ["1"] and params.get("agent") and params.get("report_id"):
    confirm_row = {
        "agent": params.get("agent")[0],
        "report_id": params.get("report_id")[0],
        "timestamp": datetime.now().isoformat(timespec="seconds"),
    }
    confirm_path = os.path.join(os.getcwd(), "confirmations.csv")
    df_confirm = pd.DataFrame([confirm_row])
    if os.path.exists(confirm_path):
        df_confirm.to_csv(confirm_path, mode="a", header=False, index=False)
    else:
        df_confirm.to_csv(confirm_path, index=False)
    st.success("Merci, votre confirmation a été enregistrée.")

with st.sidebar:
    st.header("⚙️ Paramètres")
    # SMTP
    smtp_host = st.text_input("SMTP Host", value="smtp.office365.com")
    smtp_port = st.number_input("SMTP Port", value=587, step=1)
    smtp_user = st.text_input("SMTP User (expéditeur)", value="", placeholder="prenom.nom@leadway.com")
    smtp_pass = st.text_input("SMTP Password / App Password", value="", type="password")
    bounce_email = st.text_input(
        "Adresse retour (bounces) - optionnel",
        value="",
        help="Si vide, utilise l'expediteur SMTP.",
    )
    confirm_base_url = st.text_input(
        "URL de confirmation (optionnel)",
        value="",
        help="Ex: https://ton-app.streamlit.app. Un lien 'J'ai reçu' sera ajoute au mail.",
    )
    test_to   = st.text_input("Email de test (DRY RUN)", value="y-kouadio@leadway.com")
    dry_run   = st.toggle("Mode TEST (DRY_RUN)", value=True,
                          help="Si activé, tous les mails partent vers l'adresse de test ci-dessus.")
    from_email = smtp_user or "no-reply@leadway.com"

    st.divider()
    st.subheader("⏱️ Cadence d’envoi")
    batch_size = st.slider("Taille d’une vague (agents)", min_value=10, max_value=500, value=100, step=10)
    pause_duration = st.slider("Pause entre vagues (sec.)", min_value=0, max_value=600, value=120, step=10)
    per_email_delay = st.slider("Délai par email (sec.)", min_value=0.0, max_value=2.0, value=0.8, step=0.1)
    max_retries = st.slider("Retries par email", min_value=1, max_value=5, value=3)
    fast_mode = st.checkbox("Mode rapide (réduit les pauses)", value=False)
    adaptive_mode = st.checkbox("Mode adaptatif (ralentit en cas d'erreurs SMTP)", value=True)
    ui_minimal = st.checkbox("UI minimale (réduit les erreurs d'affichage)", value=True)
    ui_every = st.number_input("MAJ UI tous les N emails", min_value=1, max_value=200, value=20, step=1)

    st.divider()
    st.subheader("💾 Reprise")
    resume_json = st.file_uploader("Reprendre depuis un fichier pending_agents_*.json (optionnel)", type=["json"])
    st.caption("Si fourni, l’envoi reprend sur cette liste d’agents restante.")

st.info("**Étape 1.** Choisir la source de données (Excel déjà prêt `data_to_generate.xlsx`).")

colA, colB = st.columns(2)
with colA:
    use_default_path = st.toggle("Utiliser le fichier par défaut (./data/data_to_generate.xlsx)", value=True)
with colB:
    uploaded_excel = st.file_uploader("Ou importe un Excel", type=["xlsx"])

# Lecture DataFrame
df_report = None
errors = []

def load_dataframe():
    if uploaded_excel is not None:
        try:
            return pd.read_excel(uploaded_excel, sheet_name="data_to_generate", dtype="object")
        except Exception:
            uploaded_excel.seek(0)
            return pd.read_excel(uploaded_excel, dtype="object")
    else:
        if not use_default_path:
            return None
        default_path = os.path.join(os.getcwd(), "data", "data_to_generate.xlsx")
        if not os.path.exists(default_path):
            st.error(f"Fichier introuvable : {default_path}")
            return None
        try:
            return pd.read_excel(default_path, sheet_name="data_to_generate", dtype="object")
        except Exception:
            return pd.read_excel(default_path, dtype="object")

def _compute_mois_label(df_agent: pd.DataFrame) -> str:
    mois_label = "Decembre 2025"
    if "DATE ECHEANCE" in df_agent.columns:
        dates = pd.to_datetime(df_agent["DATE ECHEANCE"], errors="coerce", dayfirst=True)
        if dates.notna().any():
            dt_ref = dates.dropna().max()
            mois_fr = [
                "Janvier",
                "Fevrier",
                "Mars",
                "Avril",
                "Mai",
                "Juin",
                "Juillet",
                "Aout",
                "Septembre",
                "Octobre",
                "Novembre",
                "Decembre",
            ]
            mois_label = f"{mois_fr[dt_ref.month - 1]} {dt_ref.year}"
    return mois_label


def _normalize_role(val: object) -> str:
    v = str(val).strip().lower()
    if v in {"agent", "agents"}:
        return "AGENT"
    if v in {"manager", "mgr"}:
        return "MANAGER"
    if v in {"supermanager", "super manager", "super_manager"}:
        return "SUPERMANAGER"
    if v in {"0", "", "nan", "none"}:
        return "DIRECT"
    return v.upper()


def _compute_commission_map(df_all: pd.DataFrame) -> dict:
    required = {"PRIME NETTE", "POOL TPV", "ROLE", "CODE TEAM", "EMAIL AGENT"}
    if not required.issubset(set(df_all.columns)):
        return {}

    dfc = df_all.copy()
    dfc["ROLE_NORM"] = dfc["ROLE"].apply(_normalize_role)
    dfc["EMAIL_NORM"] = dfc["EMAIL AGENT"].astype(str).str.strip().str.lower()
    dfc["CODE_TEAM"] = dfc["CODE TEAM"].astype(str).str.strip()
    dfc["POOL_FLAG"] = pd.to_numeric(dfc["POOL TPV"], errors="coerce").fillna(0)
    dfc["PRIME_NETTE_NUM"] = pd.to_numeric(dfc["PRIME NETTE"], errors="coerce").fillna(0)

    managers = (
        dfc[dfc["ROLE_NORM"] == "MANAGER"]
        .groupby("CODE_TEAM")["EMAIL_NORM"]
        .apply(lambda s: sorted(set([x for x in s if x and x != "nan"])))
        .to_dict()
    )
    supers = (
        dfc[dfc["ROLE_NORM"] == "SUPERMANAGER"]
        .groupby("CODE_TEAM")["EMAIL_NORM"]
        .apply(lambda s: sorted(set([x for x in s if x and x != "nan"])))
        .to_dict()
    )

    commissions: dict[str, float] = {}

    def add_comm(email: str, amount: float) -> None:
        if not email or email == "nan":
            return
        commissions[email] = commissions.get(email, 0.0) + float(amount)

    for row in dfc.itertuples(index=False):
        role = row.ROLE_NORM
        if role == "DIRECT":
            continue
        pool = int(row.POOL_FLAG) == 1
        prime = float(row.PRIME_NETTE_NUM)
        code = row.CODE_TEAM
        email = row.EMAIL_NORM

        if role == "AGENT":
            agent_rate = 0.07 if pool else 0.14
            add_comm(email, prime * agent_rate)

            mgrs = managers.get(code, [])
            for m in mgrs:
                add_comm(m, prime * 0.02)

            sups = supers.get(code, [])
            for s in sups:
                add_comm(s, prime * 0.01)

        elif role == "MANAGER":
            rate = 0.07 if pool else 0.14
            add_comm(email, prime * rate)
        elif role == "SUPERMANAGER":
            rate = 0.10 if pool else 0.18
            add_comm(email, prime * rate)

    return commissions


def _compute_commission_totals(df_all: pd.DataFrame) -> dict:
    required = {"PRIME NETTE", "POOL TPV", "ROLE", "CODE TEAM", "EMAIL AGENT"}
    if not required.issubset(set(df_all.columns)):
        return {"agent": 0.0, "manager": 0.0, "supermanager": 0.0, "total": 0.0}

    dfc = df_all.copy()
    dfc["ROLE_NORM"] = dfc["ROLE"].apply(_normalize_role)
    dfc["CODE_TEAM"] = dfc["CODE TEAM"].astype(str).str.strip()
    dfc["POOL_FLAG"] = pd.to_numeric(dfc["POOL TPV"], errors="coerce").fillna(0)
    dfc["PRIME_NETTE_NUM"] = pd.to_numeric(dfc["PRIME NETTE"], errors="coerce").fillna(0)

    managers = (
        dfc[dfc["ROLE_NORM"] == "MANAGER"]
        .groupby("CODE_TEAM")["ROLE_NORM"]
        .size()
        .to_dict()
    )
    supers = (
        dfc[dfc["ROLE_NORM"] == "SUPERMANAGER"]
        .groupby("CODE_TEAM")["ROLE_NORM"]
        .size()
        .to_dict()
    )

    totals = {"agent": 0.0, "manager": 0.0, "supermanager": 0.0}

    for row in dfc.itertuples(index=False):
        role = row.ROLE_NORM
        if role == "DIRECT":
            continue
        pool = int(row.POOL_FLAG) == 1
        prime = float(row.PRIME_NETTE_NUM)
        code = row.CODE_TEAM

        if role == "AGENT":
            totals["agent"] += prime * (0.07 if pool else 0.14)
            totals["manager"] += prime * 0.02 * max(1, managers.get(code, 0))
            totals["supermanager"] += prime * 0.01 * max(1, supers.get(code, 0))
        elif role == "MANAGER":
            totals["manager"] += prime * (0.07 if pool else 0.14)
        elif role == "SUPERMANAGER":
            totals["supermanager"] += prime * (0.10 if pool else 0.18)

    totals["total"] = totals["agent"] + totals["manager"] + totals["supermanager"]
    return totals


df_report = load_dataframe()
if df_report is not None:
    df_report.columns = [str(c).strip().upper() for c in df_report.columns]

    # Contrôles colonnes
    required_cols = [
        "NOM COMPLET", "PRODUIT", "PRIME TTC", "IMMATRICULATION", "TELEPHONE",
        "DUREE ( MOIS)", "DATE ECHEANCE", "LIEN DE RENOUVELLEMENT", "EMAIL AGENT"
    ]
    missing = [c for c in required_cols if c not in df_report.columns]
    if missing:
        st.warning(f"Colonnes manquantes (le PDF/Excel peut quand même sortir): {missing}")

    # Nettoyage
    if "EMAIL AGENT" not in df_report.columns:
        st.error("La colonne 'EMAIL AGENT' est obligatoire.")
        st.stop()

    df_report["EMAIL AGENT"] = df_report["EMAIL AGENT"].astype(str).str.strip().str.lower()

    # Dates + Durée (entier)
    if "DATE ECHEANCE" in df_report.columns:
        df_report["DATE ECHEANCE"] = pd.to_datetime(df_report["DATE ECHEANCE"], errors="coerce", dayfirst=True)

    duree_col = next((c for c in df_report.columns if "DUREE" in c and "MOIS" in c), None)
    if duree_col:
        df_report[duree_col] = (
            df_report[duree_col]
            .astype(str).str.replace(",", ".", regex=False)
            .replace({"": None, "nan": None, "None": None})
        )
        df_report[duree_col] = (
            pd.to_numeric(df_report[duree_col], errors="coerce").round(0).astype("Int64")
        )

    agents = df_report["EMAIL AGENT"].dropna().unique()
    st.success(f"👥 Agents uniques détectés : {len(agents)}")
    st.caption(f"🔁 Vagues prévues ≈ {ceil(len(agents)/batch_size)}")
    commission_by_agent = _compute_commission_map(df_report)
    commission_totals = _compute_commission_totals(df_report)
    st.session_state["commission_by_agent"] = commission_by_agent

    st.subheader("💰 Commissions potentielles (si tout est renouvelé)")
    c1, c2, c3, c4 = st.columns(4)
    c1.metric("Agents", f"{int(round(commission_totals['agent'])):,}".replace(",", " "))
    c2.metric("Managers", f"{int(round(commission_totals['manager'])):,}".replace(",", " "))
    c3.metric("SuperManagers", f"{int(round(commission_totals['supermanager'])):,}".replace(",", " "))
    c4.metric("Total", f"{int(round(commission_totals['total'])):,}".replace(",", " "))

    with st.expander("🔎 Aperçu des données (100 premières lignes)"):
        st.dataframe(df_report.head(100), use_container_width=True)

else:
    st.stop()

st.divider()
st.header("🚀 Envoi")

# Zone des boutons
bcol1, bcol2, bcol3 = st.columns(3)
preview_clicked = bcol1.button("Prévisualiser les destinataires")
test_clicked    = bcol2.button("Envoyer (TEST - DRY RUN)")
real_clicked    = bcol3.button("Envoyer (RÉEL)")

# Corps de mail (modifiable)
st.subheader("✉️ Message")
default_body = (
    "Bonjour,\n\n"
    "Veuillez trouver ci-joint votre liste des contrats à renouveler (PDF + Excel).\n"
    "Merci d’utiliser le lien de renouvellement pour renouveler les contrats de vos assurés "
    "et optimiser vos commissions.\n\n"
    "Cordialement,\n"
    "L’équipe Leadway Assurance."
)
email_subject = st.text_input("Sujet", value="Rapport de renouvellement – Janvier")
email_body = st.text_area("Corps du message", value=default_body, height=180)

st.subheader("🔁 Reprise automatique")
resume_from_log = st.checkbox(
    "Reprendre automatiquement les non‑envoyés (dernier log)",
    value=False,
    help="Utilise le dernier send_log_*.csv pour filtrer les agents restants.",
)

# Dossier de sortie fichiers
output_dir = os.path.join(os.getcwd(), "FILE")
os.makedirs(output_dir, exist_ok=True)

# Reprise si JSON fourni
resume_agents = None
if resume_from_log:
    try:
        logs = [f for f in os.listdir(os.getcwd()) if f.startswith("send_log_") and f.endswith(".csv")]
        if logs:
            latest = sorted(logs)[-1]
            df_log = pd.read_csv(os.path.join(os.getcwd(), latest))
            sent_agents = df_log[df_log["status"] == "sent"]["agent"].astype(str).str.lower().unique().tolist()
            resume_agents = [a for a in df_report["EMAIL AGENT"].dropna().unique().tolist() if a.lower() not in sent_agents]
            st.info(f"Reprise automatique : {len(resume_agents)} agents restants (log {latest}).")
        else:
            st.warning("Aucun send_log_*.csv trouvé pour la reprise automatique.")
    except Exception as e:
        st.error(f"Impossible de lire le dernier log : {e}")
if resume_json is not None:
    try:
        resume_agents = json.loads(resume_json.getvalue().decode("utf-8"))
        resume_agents = [a.strip().lower() for a in resume_agents if a]
        st.info(f"Reprise : {len(resume_agents)} agents chargés depuis le JSON.")
    except Exception as e:
        st.error(f"Impossible de lire le JSON de reprise : {e}")



def generate_pdf_xlsx_for_agent(agent_email: str):
    """Génère le PDF et l'Excel pour un agent; retourne (pdf_path, xlsx_path, mois_label, commission)."""
    df_agent = df_report[df_report["EMAIL AGENT"] == agent_email].reset_index(drop=True)
    if df_agent.empty:
        return None, None, None, None

    mois_label = _compute_mois_label(df_agent)
    commission_total = st.session_state.get("commission_by_agent", {}).get(agent_email, 0.0)

    # PDF : date au format JJ/MM/AAAA
    df_pdf = df_agent.copy()
    if "DATE ECHEANCE" in df_pdf.columns:
        df_pdf["DATE ECHEANCE"] = pd.to_datetime(
            df_pdf["DATE ECHEANCE"], errors="coerce", dayfirst=True
        ).dt.strftime("%d/%m/%Y")

    # Excel : natif
    df_xlsx = df_agent.copy()

    safe_base = re.sub(r'[<>:"/\\|?@ ]', "_", f"renouvellement_{agent_email}")
    pdf_path  = os.path.join(output_dir, f"{safe_base}.pdf")
    xlsx_path = os.path.join(output_dir, f"{safe_base}.xlsx")

    # PDF
    helper.report_generator(df_pdf, pdf_path, mois_label=mois_label, commission_total=commission_total)
    # XLSX
    with pd.ExcelWriter(xlsx_path, engine="openpyxl") as w:
        df_xlsx.to_excel(w, index=False, sheet_name="Contrats")
    style_excel(xlsx_path)

    return pdf_path, xlsx_path, mois_label, commission_total


def run_send(dry_run_flag: bool):
    if not smtp_user or not smtp_pass:
        st.error("Renseigne **SMTP User** et **SMTP Password** dans la barre latérale.")
        return

    # Liste d'agents à traiter
    if resume_agents is not None and len(resume_agents) > 0:
        agents_list = resume_agents
    else:
        agents_list = df_report["EMAIL AGENT"].dropna().unique().tolist()

    if not agents_list:
        st.success("Aucun agent à traiter.")
        return

    eff_pause = 0 if fast_mode else pause_duration
    eff_delay = 0 if fast_mode else per_email_delay
    base_pause = eff_pause
    base_delay = eff_delay
    smtp_errors_in_row = 0
    st.write(f"Envoi à **{len(agents_list)}** agents. Batch={batch_size}, Pause={eff_pause}s, DRY_RUN={dry_run_flag}")
    progress = st.progress(0)
    log_area = st.empty()
    recap_placeholder = st.empty()

    # SMTP
    try:
        server = open_smtp(smtp_host, smtp_port, smtp_user, smtp_pass, log=st.write)
    except Exception as e:
        st.error(f"Connexion SMTP impossible : {e}")
        return

    sent_total = 0
    failed_global = []
    log_entries = []
    timestamp = datetime.now().strftime("%Y%m%d-%H%M%S")
    log_path = os.path.join(os.getcwd(), f"send_log_{timestamp}.csv")
    st.session_state["send_log_path"] = log_path

    def append_log(entry: dict) -> None:
        log_entries.append(entry)
        file_exists = os.path.exists(log_path)
        with open(log_path, "a", newline="", encoding="utf-8") as f:
            writer = csv.DictWriter(f, fieldnames=list(entry.keys()))
            if not file_exists:
                writer.writeheader()
            writer.writerow(entry)
    total_batches = ceil(len(agents_list) / batch_size)
    ui_counter = 0
    last_ui_msg = ""
    summary_interval = max(1, int(ui_every))

    try:
        for b in range(0, len(agents_list), batch_size):
            batch_idx = b // batch_size + 1
            batch_agents = agents_list[b:b+batch_size]

            if not ui_minimal:
                st.write(f"---")
                st.write(f"**Vague {batch_idx}/{total_batches}** – {len(batch_agents)} agents")

            for i, agent in enumerate(batch_agents, start=1):
                # Génération fichiers
                try:
                    pdf_path, xlsx_path, mois_label, commission_total = generate_pdf_xlsx_for_agent(agent)
                    if not pdf_path or not xlsx_path:
                        st.warning(f"Aucun contrat pour {agent}, on saute.")
                        continue
                except Exception as ge:
                    st.error(f"❌ Génération pour {agent} : {ge}")
                    failed_global.append(agent)
                    append_log({
                        "agent": agent,
                        "status": "generation_failed",
                        "attempts": 0,
                        "subject": "",
                        "commission": "",
                        "report_id": "",
                        "confirm_url": "",
                        "pdf": "",
                        "xlsx": "",
                        "error": str(ge),
                        "timestamp": datetime.now().isoformat(timespec="seconds"),
                    })
                    continue

                # Envoi (avec petites tentatives)
                commission_label = f"{int(round(commission_total or 0)):,}".replace(",", " ")
                subject = f"Rapport de renouvellement – {mois_label or ''}".strip()

                body = email_body
                commission_line = f"<b><font color='red'>Commission potentielle si tout est renouvelé : {commission_label} FCFA.</font></b>"
                report_id = f"{agent}-{uuid.uuid4().hex[:8]}"
                confirm_url = ""
                if confirm_base_url:
                    base = confirm_base_url.rstrip("/")
                    confirm_url = (
                        f"{base}/?confirm=1&agent={quote_plus(agent)}&report_id={quote_plus(report_id)}"
                    )
                if "{COMMISSION}" in body:
                    body = body.replace("{COMMISSION}", f"<b><font color='red'>{commission_label} FCFA</font></b>")
                elif "Cordialement" in body:
                    before, sep, after = body.partition("Cordialement")
                    link_line = ""
                    if confirm_url:
                        link_line = f"<br>Confirmation : <a href=\"{confirm_url}\">J'ai reçu ce rapport</a><br>"
                    body = f"{before.rstrip()}\n{commission_line}{link_line}\n{sep}{after}"
                else:
                    link_line = f"\nConfirmation : {confirm_url}\n" if confirm_url else "\n"
                    body = body.rstrip() + f"\n{commission_line}{link_line}"
                attempt = 0
                last_error = ""
                while attempt < max_retries:
                    try:
                        try:
                            server.noop()
                        except Exception:
                            st.write("🔌 SMTP déconnecté, reconnexion…")
                            time.sleep(3)
                            server = open_smtp(smtp_host, smtp_port, smtp_user, smtp_pass, log=st.write)

                        refused = send_email_with_attachments(
                            server, from_email, agent, subject, body,
                            [pdf_path, xlsx_path], dry_run_flag, test_to,
                            request_receipt=True,
                            envelope_from=bounce_email or from_email,
                            report_id=report_id,
                        )
                        if refused:
                            raise RuntimeError(f"Refus SMTP: {refused}")
                        dest = test_to if dry_run_flag else agent
                        last_ui_msg = f"📧 Vague {batch_idx} – {i}/{len(batch_agents)} envoyé à {dest}"
                        ui_counter += 1
                        if not ui_minimal and (ui_counter % ui_every == 0):
                            log_area.info(last_ui_msg)
                        if ui_counter % summary_interval == 0:
                            recap_placeholder.success(
                                f"✅ Envoyés: {sent_total} | ❌ Échecs: {len(failed_global)}"
                            )
                        sent_total += 1
                        if adaptive_mode:
                            smtp_errors_in_row = max(0, smtp_errors_in_row - 1)
                            eff_delay = max(base_delay, eff_delay - 0.1)
                            eff_pause = max(base_pause, eff_pause - 5)
                        append_log({
                            "agent": agent,
                            "status": "sent",
                            "attempts": attempt + 1,
                            "subject": subject,
                            "commission": commission_label,
                            "report_id": report_id,
                            "confirm_url": confirm_url,
                            "pdf": pdf_path,
                            "xlsx": xlsx_path,
                            "timestamp": datetime.now().isoformat(timespec="seconds"),
                        })
                        break
                    except Exception as e:
                        attempt += 1
                        last_error = str(e)
                        if adaptive_mode:
                            smtp_errors_in_row += 1
                            eff_delay = min(5.0, eff_delay + 0.3)
                            eff_pause = min(600, eff_pause + 15)
                        last_ui_msg = f"⚠️ Tentative {attempt}/{max_retries} pour {agent} : {e}"
                        ui_counter += 1
                        if not ui_minimal and (ui_counter % ui_every == 0):
                            log_area.warning(last_ui_msg)
                        if ui_counter % summary_interval == 0:
                            recap_placeholder.info(
                                f"✅ Envoyés: {sent_total} | ❌ Échecs: {len(failed_global)}"
                            )
                        time.sleep(3)
                else:
                    st.error(f"⛔ Abandon pour {agent} après {max_retries} tentatives.")
                    failed_global.append(agent)
                    append_log({
                        "agent": agent,
                        "status": "failed",
                        "attempts": max_retries,
                        "subject": subject,
                        "commission": commission_label,
                        "pdf": pdf_path,
                        "xlsx": xlsx_path,
                        "error": last_error or "max_retries",
                        "timestamp": datetime.now().isoformat(timespec="seconds"),
                    })

                # petit throttle
                if eff_delay > 0:
                    time.sleep(eff_delay)

                if not ui_minimal and (ui_counter % ui_every == 0):
                    progress.progress(min(1.0, sent_total / max(1, len(agents_list))))

            # pause entre vagues
            if b + batch_size < len(agents_list) and eff_pause > 0:
                if not ui_minimal:
                    st.info(f"⏸️ Pause {eff_pause}s avant la prochaine vague…")
                time.sleep(eff_pause)

        st.success(f"✅ Terminé ! Emails envoyés : {sent_total}/{len(agents_list)}")
        if log_entries:
            st.info(f"Journal d'envoi : {log_path}")
            st.session_state["send_log"] = pd.DataFrame(log_entries)
        if failed_global:
            fail_path = os.path.join(os.getcwd(), f"failed_agents_{timestamp}.xlsx")
            pd.DataFrame({"EMAIL_AGENT_NON_ENVOYE": failed_global}).to_excel(fail_path, index=False)
            pending_path = os.path.join(os.getcwd(), f"pending_agents_{timestamp}.json")
            with open(pending_path, "w", encoding="utf-8") as f:
                json.dump(failed_global, f, ensure_ascii=False, indent=2)
            st.warning(f"Certains envois ont échoué. Fichiers : {fail_path}, {pending_path}")

    except KeyboardInterrupt:
        # Sauvegarder la file restante
        remaining = agents_list[sent_total:]
        fn = f"pending_agents_{timestamp}.json"
        with open(fn, "w", encoding="utf-8") as f:
            json.dump(remaining, f, ensure_ascii=False, indent=2)
        st.error(f"🛑 Interruption. Restants sauvegardés : {fn}")
    finally:
        try:
            server.quit()
        except Exception:
            pass


# ----------------------------- Actions des boutons ----------------------------

if preview_clicked:
    st.subheader("Destinataires (aperçu)")
    st.dataframe(pd.DataFrame({"EMAIL AGENT": df_report["EMAIL AGENT"].dropna().unique()}).head(200),
                 use_container_width=True)

if test_clicked:
    run_send(dry_run_flag=True)

if real_clicked:
    if dry_run:
        st.warning("Désactive 'Mode TEST (DRY_RUN)' dans la barre latérale pour un envoi RÉEL.")
    else:
        run_send(dry_run_flag=False)

st.divider()
st.header("📎 Envoi manuel de fichiers aux agents")

manual_subject = st.text_input("Sujet (envoi manuel)", value="BNB AUX AGENTS")
manual_body = st.text_area(
    "Message (envoi manuel)",
    value=(
        "Cher partenaire,\n\n"
        "vous avez possibilité d’épargner par mois 9.450 fcfa et de gagner 500.000 fcfa "
        "par tirage au sort ou à la fin du contrat.\n\n"
        "Le bulletin d’adhésion et la fiche d’autorisation de prélèvement sur les commissions "
        "hebdomadaires sont à renseigner et à ramener par mail à g-beda@leadway.com "
        "ou r-anougre@leadway.com\n\n"
        "Cordialement,"
    ),
    height=180,
)
manual_files = st.file_uploader(
    "Pièces jointes (Excel, Word, PDF)",
    type=["xlsx", "xls", "docx", "doc", "pdf"],
    accept_multiple_files=True,
)
send_to_all = st.checkbox("Envoyer à tous les agents", value=False)
manual_list = st.text_area(
    "Ou liste manuelle d'emails (un par ligne)",
    value="",
    height=100,
)

if send_to_all:
    manual_recipients = df_report["EMAIL AGENT"].dropna().unique().tolist()
else:
    manual_recipients = [e.strip().lower() for e in manual_list.splitlines() if e.strip()]

if st.button("📤 Envoyer (manuel)"):
    if not smtp_user or not smtp_pass:
        st.error("Renseigne **SMTP User** et **SMTP Password** dans la barre latérale.")
    elif not manual_recipients:
        st.error("Aucun destinataire.")
    elif not manual_files:
        st.error("Ajoute au moins un fichier.")
    else:
        try:
            server = open_smtp(smtp_host, smtp_port, smtp_user, smtp_pass, log=st.write)
        except Exception as e:
            st.error(f"Connexion SMTP impossible : {e}")
            server = None
        if server:
            sent = 0
            failed = 0
            tmp_dir = os.path.join(os.getcwd(), "manual_attachments")
            os.makedirs(tmp_dir, exist_ok=True)
            paths = []
            for f in manual_files:
                p = os.path.join(tmp_dir, f.name)
                with open(p, "wb") as out:
                    out.write(f.read())
                paths.append(p)
            for dest in manual_recipients:
                try:
                    try:
                        server.noop()
                    except Exception:
                        server = open_smtp(smtp_host, smtp_port, smtp_user, smtp_pass, log=st.write)
                    send_email_with_attachments(
                        server,
                        smtp_user,
                        dest,
                        manual_subject,
                        manual_body,
                        paths,
                        dry_run=False,
                        test_to="",
                    )
                    sent += 1
                except Exception as e:
                    failed += 1
                    st.warning(f"Erreur envoi à {dest} : {e}")
            try:
                server.quit()
            except Exception:
                pass
            st.success(f"Envoi terminé. Envoyés: {sent}, Échecs: {failed}.")

send_log_df = st.session_state.get("send_log")
log_path_hint = st.session_state.get("send_log_path")
if send_log_df is None:
    try:
        logs = [f for f in os.listdir(os.getcwd()) if f.startswith("send_log_") and f.endswith(".csv")]
        if logs:
            latest = sorted(logs)[-1]
            send_log_df = pd.read_csv(os.path.join(os.getcwd(), latest))
    except Exception:
        send_log_df = None

if send_log_df is not None and not send_log_df.empty:
    st.subheader("📋 Suivi des envois")
    st.dataframe(send_log_df, use_container_width=True)
    counts = send_log_df["status"].value_counts()
    col_a, col_b, col_c = st.columns(3)
    col_a.metric("Envoyés", int(counts.get("sent", 0)))
    col_b.metric("Échecs", int(counts.get("failed", 0)))
    col_c.metric("Génération KO", int(counts.get("generation_failed", 0)))
    if log_path_hint and os.path.exists(log_path_hint):
        try:
            with open(log_path_hint, "rb") as f:
                st.download_button(
                    "⬇️ Télécharger le dernier log",
                    data=f.read(),
                    file_name=os.path.basename(log_path_hint),
                    mime="text/csv",
                )
        except Exception:
            pass
    st.download_button(
        "⬇️ Télécharger le suivi (CSV)",
        data=send_log_df.to_csv(index=False).encode("utf-8"),
        file_name=f"send_log_{datetime.now().strftime('%Y%m%d-%H%M%S')}.csv",
        mime="text/csv",
    )
elif log_path_hint:
    st.info(f"Aucun suivi chargé pour l'instant. Dernier chemin de log: {log_path_hint}")

st.subheader("📮 Analyse des retours SMTP (bounces)")
eml_files = st.file_uploader(
    "Importer des fichiers .eml de bounce/DSN",
    type=["eml"],
    accept_multiple_files=True,
)
if eml_files:
    rows = []
    for f in eml_files:
        try:
            rows.extend(parse_dsn_eml(f.read()))
        except Exception as e:
            rows.append({
                "Final-Recipient": "",
                "Action": "",
                "Status": "",
                "Diagnostic-Code": f"Erreur lecture: {e}",
                "Original-Message-ID": "",
            })
    if rows:
        st.dataframe(pd.DataFrame(rows), use_container_width=True)

if os.path.exists(os.path.join(os.getcwd(), "confirmations.csv")):
    st.subheader("✅ Confirmations de lecture (lien)")
    try:
        st.dataframe(pd.read_csv(os.path.join(os.getcwd(), "confirmations.csv")), use_container_width=True)
    except Exception:
        pass
