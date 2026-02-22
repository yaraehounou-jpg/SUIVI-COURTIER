# =============================================================================
# main.py – Envoi des rapports de renouvellement par agent (PDF + Excel)
# =============================================================================
# -*- coding: utf-8 -*-
import os
import re
import json
import time
import ssl
import smtplib
import mimetypes
import pandas as pd
from math import ceil
from datetime import datetime
from email.message import EmailMessage

from helper import Helper  # utilise helper.report_generator
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# =============================================================================
# 1) CONFIG SMTP
# =============================================================================
SMTP_HOST = "smtp.office365.com"
SMTP_PORT = 587
SMTP_USER = "leadvie@leadway.com"      # <-- A RENSEIGNER
SMTP_PASS = "bM_BS[EZc7/m"        # <-- A RENSEIGNER (mot de passe d’application recommandé)

FROM_EMAIL = SMTP_USER
DRY_RUN = False                             # True = tout va vers TEST_TO
TEST_TO = "y-kouadio@leadway.com"

# Vagues & rythme
batch_size = 100                           # nb agents par vague
pause_duration = 120                       # pause (secondes) entre vagues
PER_EMAIL_DELAY = 0.8                      # throttle léger par email (sec)
MAX_RETRIES_PER_EMAIL = 3                  # retries en cas d'erreur SMTP
RECONNECT_SLEEP = 5                        # attente avant reconnexion

# Reprise optionnelle (si tu veux reprendre sur une file sauvegardée)
RESUME_FILE = ""                           # ex: "pending_agents_20251008-1530.json"


# =============================================================================
# 2) OUTILS – Excel styling, SMTP, Email
# =============================================================================
def style_excel(xlsx_path: str, sheet_name: str = "Contrats", date_col_name: str = "Date echeance"):
    """Applique un style simple et lisible sur le fichier Excel."""
    wb = load_workbook(xlsx_path)
    ws = wb[sheet_name]

    max_row, max_col = ws.max_row, ws.max_column

    header_fill = PatternFill(start_color="FF8C00", end_color="FF8C00", fill_type="solid")
    header_font = Font(name="Helvetica", bold=True, color="FFFFFF", size=10)
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for j in range(1, max_col + 1):
        c = ws.cell(row=1, column=j)
        c.fill, c.font, c.alignment = header_fill, header_font, header_align

    thin = Side(style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    body_font = Font(name="Helvetica", size=9)
    body_align = Alignment(horizontal="left", vertical="top", wrap_text=True)

    for i in range(2, max_row + 1):
        for j in range(1, max_col + 1):
            c = ws.cell(row=i, column=j)
            c.border, c.font, c.alignment = border, body_font, body_align

    # largeur auto basique (en fonction de l'entête)
    for j in range(1, max_col + 1):
        header_text = str(ws.cell(1, j).value or "")
        base = max(12, min(60, int(len(header_text) * 1.2) + 2))
        ws.column_dimensions[get_column_letter(j)].width = base

    # format date si présent
    try:
        date_idx = next((j for j in range(1, max_col + 1)
                         if str(ws.cell(1, j).value).strip() == date_col_name), None)
        if date_idx:
            for r in range(2, max_row + 1):
                ws.cell(r, date_idx).number_format = "DD/MM/YYYY"
    except Exception:
        pass

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    wb.save(xlsx_path)


def open_smtp():
    """Ouvre une connexion SMTP Office365 avec TLS."""
    ctx = ssl.create_default_context()
    server = smtplib.SMTP(SMTP_HOST, SMTP_PORT, timeout=60)
    server.ehlo()
    server.starttls(context=ctx)
    server.ehlo()
    server.login(SMTP_USER, SMTP_PASS)
    print("📡 Connexion SMTP OK")
    return server


def send_email_with_attachments(server, to_email, subject, body, attachments: list):
    msg = EmailMessage()
    msg["Subject"] = subject
    msg["From"] = f"Leadway Renouvellements <{FROM_EMAIL}>"
    msg["To"] = TEST_TO if DRY_RUN else to_email
    msg["Reply-To"] = FROM_EMAIL
    msg.set_content(body)

    for path in attachments:
        if not os.path.exists(path):
            print(f"⚠️ PJ introuvable : {path}")
            continue
        ctype, _ = mimetypes.guess_type(path)
        if ctype is None:
            ctype = "application/octet-stream"
        maintype, subtype = ctype.split("/", 1)
        with open(path, "rb") as f:
            msg.add_attachment(
                f.read(),
                maintype=maintype,
                subtype=subtype,
                filename=os.path.basename(path),
            )

    server.send_message(msg)


# =============================================================================
# 3) CHEMINS & LECTURE DES DONNÉES
# =============================================================================
script_dir = os.path.dirname(os.path.abspath(__file__))
input_path = os.path.join(script_dir, "data", "data_to_generate.xlsx")
output_dir = os.path.join(script_dir, "FILE")
os.makedirs(output_dir, exist_ok=True)

# Lecture : on tente la feuille 'data_to_generate', sinon la 1re feuille
if not os.path.exists(input_path):
    raise FileNotFoundError(f"Fichier introuvable : {input_path}")

try:
    df_report = pd.read_excel(input_path, sheet_name="data_to_generate", dtype="object")
except Exception:
    df_report = pd.read_excel(input_path, dtype="object")

# Normalisation des noms de colonnes (insensible à la casse / espaces)
df_report.columns = [str(c).strip().upper() for c in df_report.columns]

# Colonnes attendues minimales
required_cols = [
    "NOM COMPLET", "PRODUIT", "PRIME TTC", "IMMATRICULATION", "TELEPHONE",
    "DUREE ( MOIS)", "DATE ECHEANCE", "LIEN DE RENOUVELLEMENT", "EMAIL AGENT"
]
missing = [c for c in required_cols if c not in df_report.columns]
if missing:
    print(f"⚠️ Colonnes manquantes dans le fichier Excel : {missing}")

# Nettoyage de l'email agent
if "EMAIL AGENT" not in df_report.columns:
    raise KeyError("La colonne 'EMAIL AGENT' est obligatoire dans data_to_generate.xlsx.")

df_report["EMAIL AGENT"] = (
    df_report["EMAIL AGENT"].astype(str).str.strip().str.lower()
)

def _normalize_role(val):
    v = str(val).strip().lower()
    if v in {"agent", "agents"}:
        return "AGENT"
    if v in {"manager", "mgr"}:
        return "MANAGER"
    if v in {"supermanager", "super manager", "super_manager"}:
        return "SUPERMANAGER"
    if v in {"0", "", "nan", "none"}:
        return "DIRECT"
    return v.upper()


def _compute_commission_map(df_all):
    required = {"PRIME NETTE", "POOL TPV", "ROLE", "CODE TEAM", "EMAIL AGENT"}
    if not required.issubset(set(df_all.columns)):
        return {}

    dfc = df_all.copy()
    dfc["ROLE_NORM"] = dfc["ROLE"].apply(_normalize_role)
    dfc["EMAIL_NORM"] = dfc["EMAIL AGENT"].astype(str).str.strip().str.lower()
    dfc["CODE_TEAM"] = dfc["CODE TEAM"].astype(str).str.strip()
    dfc["POOL_FLAG"] = pd.to_numeric(dfc["POOL TPV"], errors="coerce").fillna(0)
    dfc["PRIME_NETTE_NUM"] = pd.to_numeric(dfc["PRIME NETTE"], errors="coerce").fillna(0)

    managers = (
        dfc[dfc["ROLE_NORM"] == "MANAGER"]
        .groupby("CODE_TEAM")["EMAIL_NORM"]
        .apply(lambda s: sorted(set([x for x in s if x and x != "nan"])))
        .to_dict()
    )
    supers = (
        dfc[dfc["ROLE_NORM"] == "SUPERMANAGER"]
        .groupby("CODE_TEAM")["EMAIL_NORM"]
        .apply(lambda s: sorted(set([x for x in s if x and x != "nan"])))
        .to_dict()
    )

    commissions = {}

    def add_comm(email, amount):
        if not email or email == "nan":
            return
        commissions[email] = commissions.get(email, 0.0) + float(amount)

    for row in dfc.itertuples(index=False):
        role = row.ROLE_NORM
        if role == "DIRECT":
            continue
        pool = int(row.POOL_FLAG) == 1
        prime = float(row.PRIME_NETTE_NUM)
        code = row.CODE_TEAM
        email = row.EMAIL_NORM

        if role == "AGENT":
            agent_rate = 0.07 if pool else 0.14
            add_comm(email, prime * agent_rate)

            mgrs = managers.get(code, [])
            if mgrs:
                for m in mgrs:
                    add_comm(m, prime * 0.02)

            sups = supers.get(code, [])
            if sups:
                for s in sups:
                    add_comm(s, prime * 0.01)

        elif role == "MANAGER":
            rate = 0.07 if pool else 0.14
            add_comm(email, prime * rate)
        elif role == "SUPERMANAGER":
            rate = 0.10 if pool else 0.18
            add_comm(email, prime * rate)

    return commissions

commission_by_agent = _compute_commission_map(df_report)

# Conversion date échéance (si présente)
if "DATE ECHEANCE" in df_report.columns:
    df_report["DATE ECHEANCE"] = pd.to_datetime(
        df_report["DATE ECHEANCE"], errors="coerce", dayfirst=True
    )

# Durée en entier (si présente)
duree_col = next((c for c in df_report.columns if "DUREE" in c and "MOIS" in c), None)
if duree_col:
    df_report[duree_col] = (
        df_report[duree_col]
        .astype(str).str.replace(",", ".", regex=False)
        .replace({"": None, "nan": None, "None": None})
    )
    df_report[duree_col] = (
        pd.to_numeric(df_report[duree_col], errors="coerce")
        .round(0)
        .astype("Int64")
    )

# Liste d'agents
all_agents = (
    df_report["EMAIL AGENT"].dropna().unique().tolist()
)
nb_agents = len(all_agents)

print(f"👥 Agents uniques détectés : {nb_agents}")
print(f"📦 Batch size : {batch_size} | ⏸️ Pause : {pause_duration}s")
print(f"🔁 Vagues prévues : {ceil(nb_agents / batch_size)}")


# =============================================================================
# 4) HELPER (PDF)
# =============================================================================
helper = Helper()  # helper.report_generator crée le PDF compact (logo + bandeau orange)


# =============================================================================
# 5) ENVOI PAR VAGUES – avec retries, reconnexion, sauvegarde file restante
# =============================================================================
def send_batch(agent_list, server, batch_idx):
    """Envoie une vague d'emails aux agents donnés. Retourne (failed, sent)."""
    failed = []
    sent = []
    total = len(agent_list)

    for idx, agent in enumerate(agent_list, start=1):
        df_agent = df_report[df_report["EMAIL AGENT"] == agent].reset_index(drop=True)
        if df_agent.empty:
            print(f"⚠️ Aucun contrat pour {agent}, on saute.")
            continue

        # Préparation PDF/Excel par agent
        # PDF : Date echeance en JJ/MM/AAAA
        df_pdf = df_agent.copy()
        if "DATE ECHEANCE" in df_pdf.columns:
            df_pdf["DATE ECHEANCE"] = pd.to_datetime(
                df_pdf["DATE ECHEANCE"], errors="coerce", dayfirst=True
            ).dt.strftime("%d/%m/%Y")

        # Excel : garde datetime pour format Excel
        df_xlsx = df_agent.copy()

        # Fichiers
        safe_base = re.sub(r'[<>:"/\\|?@ ]', "_", f"renouvellement_{agent}")
        pdf_path = os.path.join(output_dir, f"{safe_base}.pdf")
        xlsx_path = os.path.join(output_dir, f"{safe_base}.xlsx")

        # Génération
        try:
            mois_label = "Decembre 2025"
            if "DATE ECHEANCE" in df_agent.columns:
                dates = pd.to_datetime(df_agent["DATE ECHEANCE"], errors="coerce", dayfirst=True)
                if dates.notna().any():
                    dt_ref = dates.dropna().max()
                    mois_fr = [
                        "Janvier",
                        "Fevrier",
                        "Mars",
                        "Avril",
                        "Mai",
                        "Juin",
                        "Juillet",
                        "Aout",
                        "Septembre",
                        "Octobre",
                        "Novembre",
                        "Decembre",
                    ]
                    mois_label = f"{mois_fr[dt_ref.month - 1]} {dt_ref.year}"
            commission_total = commission_by_agent.get(agent, 0.0)
            helper.report_generator(df_pdf, pdf_path, mois_label=mois_label, commission_total=commission_total)
        except Exception as e:
            print(f"❌ Erreur génération PDF pour {agent} : {e}")
            failed.append(agent)
            continue

        try:
            with pd.ExcelWriter(xlsx_path, engine="openpyxl") as w:
                df_xlsx.to_excel(w, index=False, sheet_name="Contrats")
            style_excel(xlsx_path)
        except Exception as e:
            print(f"❌ Erreur génération Excel pour {agent} : {e}")
            failed.append(agent)
            continue

        # Email
        commission_total = commission_by_agent.get(agent, 0.0)
        commission_label = f"{int(round(commission_total)):,}".replace(",", " ")

        subject = f"Rapport de renouvellement – {mois_label}"
        body = (
            "Bonjour,\n\n"
            "Veuillez trouver ci-joint votre rapport de renouvellement (PDF + Excel).\n"
            f"Commission potentielle si tout est renouvelé : {commission_label} FCFA.\n"
            "Merci d’utiliser le lien de renouvellement pour renouveler les contrats de vos assurés "
            "et optimiser vos commissions.\n\n"
            "Cordialement,\n"
            "L’équipe Leadway Assurance."
        )

        attempt = 0
        while attempt < MAX_RETRIES_PER_EMAIL:
            try:
                # Vérifier la connexion, sinon reconnecter
                try:
                    server.noop()
                except Exception:
                    print("🔌 SMTP déconnecté, reconnexion…")
                    time.sleep(RECONNECT_SLEEP)
                    server = open_smtp()

                send_email_with_attachments(server, agent, subject, body, [pdf_path, xlsx_path])
                dest = TEST_TO if DRY_RUN else agent
                print(f"📧 (batch {batch_idx} | {idx}/{total}) envoyé à {dest}")
                sent.append(agent)
                break

            except (smtplib.SMTPServerDisconnected,
                    smtplib.SMTPDataError,
                    smtplib.SMTPRecipientsRefused) as e:
                attempt += 1
                print(f"⚠️ Envoi à {agent} échec (tentative {attempt}/{MAX_RETRIES_PER_EMAIL}) : {e}")
                time.sleep(RECONNECT_SLEEP)
                try:
                    server = open_smtp()
                except Exception as e2:
                    print(f"❌ Reconnexion SMTP échouée : {e2}")
            except Exception as e:
                attempt += 1
                print(f"❌ Erreur inattendue à {agent} (tentative {attempt}/{MAX_RETRIES_PER_EMAIL}) : {e}")
                time.sleep(RECONNECT_SLEEP)
        else:
            print(f"⛔ Abandon de l’agent {agent} après {MAX_RETRIES_PER_EMAIL} tentatives.")
            failed.append(agent)

        # petit throttle pour éviter saturation
        time.sleep(PER_EMAIL_DELAY)

    return failed, sent


# =============================================================================
# 6) LANCEMENT – Gestion interruption & reprise
# =============================================================================
def main():
    # File de travail (reprise possible)
    if RESUME_FILE and os.path.exists(RESUME_FILE):
        with open(RESUME_FILE, "r", encoding="utf-8") as f:
            agents = json.load(f)
        print(f"🔁 Reprise depuis {RESUME_FILE} – {len(agents)} agents restants.")
    else:
        agents = list(all_agents)

    if not agents:
        print("✅ Aucun agent à traiter.")
        return

    timestamp = datetime.now().strftime("%Y%m%d-%H%M%S")
    pending = list(agents)

    server = open_smtp()
    try:
        total_batches = ceil(len(agents) / batch_size)
        for b in range(0, len(agents), batch_size):
            batch_idx = b // batch_size + 1
            batch_agents = agents[b:b + batch_size]
            print(f"\n🚀 Vague {batch_idx}/{total_batches} – {len(batch_agents)} agents…")

            failed_batch, sent_batch = send_batch(batch_agents, server, batch_idx)

            # Retirer envoyés de la file
            for a in sent_batch:
                if a in pending:
                    pending.remove(a)

            # Sauver les fails de la vague (optionnel)
            if failed_batch:
                fail_path = os.path.join(script_dir, f"failed_agents_{timestamp}_b{batch_idx}.xlsx")
                pd.DataFrame({"EMAIL_AGENT_NON_ENVOYE": failed_batch}).to_excel(fail_path, index=False)
                print(f"⚠️ Envois échoués (vague {batch_idx}) → {fail_path}")

            # Pause si d'autres vagues à venir
            if b + batch_size < len(agents):
                print(f"⏸️ Pause de {pause_duration}s avant la prochaine vague…")
                time.sleep(pause_duration)

        print("\n✅ Envois terminés.")

    except KeyboardInterrupt:
        # Sauvegarder la file restante
        fn = os.path.join(script_dir, f"pending_agents_{timestamp}.json")
        with open(fn, "w", encoding="utf-8") as f:
            json.dump(pending, f, ensure_ascii=False, indent=2)
        print(f"\n🛑 Interruption manuelle. File restante sauvegardée dans : {fn}")
    finally:
        try:
            server.quit()
        except Exception:
            pass


if __name__ == "__main__":
    main()
